import openpyxl
import os

# --- パス定義 ---
# このファイルが存在するディレクトリ（/school_qna）を取得
SYSTEM_DIR = os.path.dirname(os.path.abspath(__file__))
# 管理者用_touchable フォルダへのパスを定義 (../../管理者用_touchable)
TOUCHABLE_DIR = os.path.join(SYSTEM_DIR, '..', '..', '管理者用_touchable')
CATEGORIES_FILE = os.path.join(TOUCHABLE_DIR, '質問内容と小区分の編集.xlsx')

def create_categories_template_if_not_exists():
    """
    '質問内容と小区分の編集.xlsx' が存在しない場合に、
    ヘッダー付きのテンプレートファイルを自動生成する関数。
    """
    # フォルダが存在しない場合は作成（通常は入退室アプリ側であるはずだが念のため）
    if not os.path.exists(TOUCHABLE_DIR):
        os.makedirs(TOUCHABLE_DIR, exist_ok=True)

    if not os.path.exists(CATEGORIES_FILE):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "質問項目"
            
            # ヘッダーを作成
            headers = ["質問内容"]
            for i in range(1, 101):
                headers.append(f"小区分{i}")
            
            sheet.append(headers)

            # サンプルの説明を追加
            sample_row = ["（例）数学Ⅰ", "数と式", "二次関数", "図形と計量", "データの分析"]
            sheet.append(sample_row)
            
            workbook.save(CATEGORIES_FILE)
            print(f"'{CATEGORIES_FILE}' が存在しなかったため、テンプレートを自動生成しました。")
            print("このファイルに質問内容と小区分を記入してください。")

        except Exception as e:
            print(f" カテゴリファイルのテンプレート作成中にエラーが発生しました: {e}")


def load_sub_categories():
    """
    '質問内容と小区分の編集.xlsx' を読み込み、
    質問内容と小区分の辞書を作成する関数。
    """
    # 起動時にファイルが存在するか確認し、なければテンプレートを作成
    create_categories_template_if_not_exists()

    sub_categories = {}
    
    try:
        workbook = openpyxl.load_workbook(CATEGORIES_FILE, data_only=True)
        sheet = workbook.active
        
        # 1行目はヘッダーなので、2行目から読み込む
        for row in sheet.iter_rows(min_row=2, values_only=True):
            subject = row[0]
            if not subject:
                # 質問内容のセルが空なら、その行はスキップ
                continue
            
            # 小区分をリストとして取得（空のセルは無視する）
            categories = [cell for cell in row[1:] if cell is not None and str(cell).strip() != ""]
            
            sub_categories[subject] = categories
            
        print(f"'{os.path.basename(CATEGORIES_FILE)}' から質問項目を読み込みました。")
        if not sub_categories:
             print("警告: 質問項目が1件も読み込まれませんでした。ファイルが空か、フォーマットが正しいか確認してください。")

    except FileNotFoundError:
        print(f"🚨 '{CATEGORIES_FILE}' が見つかりませんでした。空の状態で処理を続行します。")
        # テンプレート作成関数が呼ばれているので、基本的にはこのエラーは起きないはず
    except Exception as e:
        print(f"🚨 カテゴリファイルの読み込み中にエラーが発生しました: {e}")

    return sub_categories

if __name__ == '__main__':
    # このファイルを直接実行した場合のテスト用
    loaded_data = load_sub_categories()
    print("\n--- 読み込みデータ確認 ---")
    print(loaded_data)
    print("--------------------------")